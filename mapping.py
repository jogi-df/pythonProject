import openpyxl
import pandas

#now lets copy the columns from raw to the template but I only know how to do it in pandas
wbt = openpyxl.load_workbook(r'c:\test\new.xlsx')
wst = wbt['Template (File Format)']

wbf = openpyxl.load_workbook(r'c:\test\testrawdata_filtered.xlsx')
wsf = wbf['Sheet1']

wb3 = openpyxl.load_workbook(r'c:\test\col_map.xlsx')
ws3 = wb3['Sheet1']


# get map column ids
mr = ws3.max_row
mc = ws3.max_column

mr1 = wsf.max_row

for x in range (2, mr + 1):
    a = ws3.cell(row = x, column = 1).value
    b = ws3.cell(row = x, column = 2).value
    if a is None:
        continue
    for i in range (1, mr1 + 1):
        c = wsf.cell(row = i, column = a)
        wst.cell(row = i, column = b).value = c.value

wbt.save(r'c:\test\final.xlsx')

