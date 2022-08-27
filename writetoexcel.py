import openpyxl
import pandas

#now lets copy the columns from raw to the template but I only know how to do it in pandas
df_template = openpyxl.load_workbook(r'c:\test\new.xlsx')
df_filtered = openpyxl.load_workbook(r'c:\test\testrawdata_filtered.xlsx')

from openpyxl.utils import rows_from_range

def copy_range(range_str, src, dst):

    for row in rows_from_range(range_str):
        for cell in row:
            dst[cell].value = src[cell].value

    return

#wb = openpyxl.load_workbook('file1.xlsx', data_only=True)
#wb1 = openpyxl.load_workbook('file2.xlsx')
sheet = df_template['Template (File Format)']
sheet1 = df_filtered['Sheet1']
for row in sheet1['A2':'Z100']:
    for cell in row:
        sheet[cell.coordinate].value = cell.value


#print (df_template)
df_template.save(r'C:\test\testrawdata_final.xlsx')

#df_template['Campaign ID'] = df_filtered['CID'].copy()
#df_template['Permissions Create Date '] = df_filtered['Permissions Create Date'].copy()
#df_template['Email'] = df_filtered['Email'].copy()
#df_template['Zip'] = df_filtered['Zip'].copy()
#df_template['Country'] = df_filtered['Country'].copy()
#df_template['Phone'] = df_filtered['Phone'].copy()
#df_template['First Name'] = df_filtered['First Name'].copy()
#df_template['Last Name'] = df_filtered['Last Name'].copy()
#df_template['Company Name'] = df_filtered['Company Name'].copy()
#df_template['OPT IN EMAIL'] = df_filtered['OPT IN EMAIL'].copy()
#df_template['OPT IN MAIL'] = df_filtered['OPT IN MAIL'].copy()
#df_template['OPT IN PHONE'] = df_filtered['OPT IN PHONE'].copy()
#df_template['Product Interest'] = df_filtered['Product Interest'].copy()
#df_template['Industry'] = df_filtered['Industry'].copy()
#df_template = pandas.concat([df_template, df_field], axis=1)


