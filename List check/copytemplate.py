#create new workbook and rename the default sheet to Template (File Format)
from openpyxl import load_workbook, Workbook

wb = load_workbook(r"C:\test\Lead Import Template Worldwide.xlsx")


sheets = wb.sheetnames

for s in sheets:
    if s != 'Template (File Format)':
        sheet_name = wb[s]
        wb.remove(sheet_name)

wb.save(r'c:\test\new1.xlsx')
