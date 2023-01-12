import pandas as pd
import easygui
import openpyxl as op
import numpy as np
import os
from functions import check_optional_col, check_required_col, opt_in, opt_in_check, check_phone_col, whitespace_remover, check_exists, state_check, zip_city, zip_state
from validate_email import validate_email
from io import StringIO
import phonenumbers
from uszipcode import SearchEngine, SimpleZipcode

#reference files
#df_zip = pd.read_excel(r'C:\test\ZIP_Locale_Detail.xls')

# turn checks on or off
#change_perm_date = "off"
phone_remove_zeros = "off"



#sets the working folder
working_folder = easygui.diropenbox(msg="Select the working folder", title="Working folder")
fold = StringIO(working_folder)

#easygui requests a file select for the main data frame.  Easygui returns the path of the file
raw_file_path = easygui.fileopenbox(msg="Select the raw data file.  Don't use the original!", default=working_folder + "\*")
file = StringIO(raw_file_path)
df = pd.read_excel(raw_file_path)
#df = pd.read_excel(r"C:\test\testrawdata.xlsx")   #uncomment this for testing and comment out everything else

# these are from the Adobe Lead Import Template Worldwide tabs and are used to compare against
#from io import StringIO
#raw_file_path1 = easygui.fileopenbox(msg="Select the template data file.  Don't use the original!")
#file = StringIO(raw_file_path1)
#df_acceptable = pd.read_excel(raw_file_path1, sheet_name="Acceptable List of Values")
#df_states = pd.read_excel(raw_file_path1, sheet_name="Territory State List")
df_acceptable = pd.read_excel(r'C:\test\Lead Import Template Worldwide.xlsx', sheet_name="Acceptable List of Values")          #uncomment this for testing and comment out everything else
df_states = pd.read_excel(r'C:\test\Lead Import Template Worldwide.xlsx', sheet_name="Territory State List")                   #uncomment this for testing and comment out everything else

#remove any trailing whitespaces in the column names (shows up in permissions create date)
df.columns = df.columns.str.rstrip()

#whitespace_remover(df)

#print(df.columns)

#inefficient column name check section--could probably loop this

if not 'Campaign ID' in df.columns:                                     #renames the column header if it is non standard
    x = [col for col in df.columns if "Campaign" in col]
    y = x[0]
    df.rename(columns={y:'Campaign ID'}, inplace=True)

if not 'Campaign ID' in df.columns:                                     #renames the column header if it is non standard
    x = [col for col in df.columns if "CID" in col]
    y = x[0]
    df.rename(columns={y:'Campaign ID'}, inplace=True)

if not 'First Name' in df.columns:
    x = [col for col in df.columns if "First" in col]
    y = x[0]
    df.rename(columns={y:'First Name'}, inplace=True)

if not 'Last Name' in df.columns:
    x = [col for col in df.columns if "Last" in col]
    y = x[0]
    df.rename(columns={y:'Last Name'}, inplace=True)

if not 'State' in df.columns:
    df['State'] = df.filter(like="State")

if not 'Zip' in df.columns:
    df['Zip'] = df.filter(like="Zip")

if not 'Email' in df.columns:
    df['Email'] = df.filter(like="Email")

if not 'Country' in df.columns:
    df['Country'] = df.filter(like="Country")

if not 'Company' in df.columns:
    x = [col for col in df.columns if "Company" in col]
    y = x[0]
    df.rename(columns={y:'Company Name'}, inplace=True)

if not 'Employee Range' in df.columns:
    df['Employee Range'] = df.filter(like="Employee")


#replace these values explicitly
#fix est number of units if wrong values
#df.loc[df['Estimated Number of Units'] == '1', 'Estimated Number of Units'] = '1'
df.loc[df['Timeline for Purchasing'] == '1 to 3 months', 'Timeline for Purchasing'] = '1-3 months'
df.loc[df['Timeline for Purchasing'] == '4 to 6 months', 'Timeline for Purchasing'] = '4-6 months'
df.loc[df['Timeline for Purchasing'] == '7 to 9 months', 'Timeline for Purchasing'] = '7-9 months'
df.loc[df['Timeline for Purchasing'] == '10 to 12 months', 'Timeline for Purchasing'] = '10-12 months'
df.loc[df['Timeline for Purchasing'] == '10 To 12 Months', 'Timeline for Purchasing'] = '10-12 months'
df.loc[df['Industry'] == 'Advertising ', 'Industry'] = 'Advertising'
df.loc[df['Country'] == 'US', 'Country'] = 'United States'
df.loc[df['Attended'] == 'Registered', 'Attended'] = 'Responded/Registered'
df.loc[df['Employee Range'] == '50-99', 'Employee Range'] = '10-99'

#remove leading and trailing spaces
df['Industry'] = df['Industry'].str.replace(r"^ +| +$", r"", regex=True)
df['Employee Range'] = df['Employee Range'].str.replace(r"^ +| +$", r"", regex=True)
df['Email'] = df['Email'].str.replace(r"^ +| +$", r"", regex=True)



#check CID length for 18 characters
df.loc[df['Campaign ID'].apply(len) == 18, 'CID_status'] = 'TRUE'
df.loc[df['Campaign ID'].apply(len) != 18, 'CID_status'] = 'FALSE'

# date section ------------------------------------------------------------------------------


#print(df['Permissions Create Date'].dtypes)

if df['Permissions Create Date'].dtypes != 'object':
    df['Permissions Create Date'] = df['Permissions Create Date'].astype(str)

#if df['Permissions Create Date'].len()  8:
if df['Permissions Create Date'].str.contains("\/|-|\.").any():     #checks to see if non numberic characters exist in date (wrong format?)
    df["Permissions Create Date"] = pd.to_datetime(df["Permissions Create Date"]).dt.strftime("%m%d%Y")

#if change_perm_date == 'on':
#    df["Permissions Create Date"] = pd.to_datetime(df["Permissions Create Date"]).dt.strftime("%m%d%Y")

df['Permissions Create Date'] = df['Permissions Create Date'].str.zfill(8)
df['perm'] = df["Permissions Create Date"].str.match("^[']?[0-1][0-9][0-3][0-9](202)[2-3]") #checks date format

#print(df['Permissions Create Date'].dtypes)


check_optional_col('Attended','Attend',df,df_acceptable)

#check if email address is in a standard format
df['Email_format'] = df['Email'].apply(validate_email)
df['Email unique'] = ~df['Email'].duplicated(keep=False)

check_optional_col('Salutation','Sal Good',df,df_acceptable)

#check_exists('First Name', 'F Name',df)
df['First Name'] = df['First Name'].replace(r"^ +| +$", r"", regex=True)
df['First Name'] = df['First Name'].str.title()
df['F Name'] = df['First Name'].str.match("^[_A-z|(-|'|á|á|ć|é|ñ|ó|è|Ô|ç|í)?]*((\s)*[_A-z])*$")

#check_exists('Last Name', 'L Name', df)
df['Last Name'] = df['Last Name'].replace(r"^ +| +$", r"", regex=True)          #remove leading and trailing spaces
df['Last Name'] = df['Last Name'].replace(r",\W*\b[A-Za-z].+\b(?!'\b)\.?", r"", regex=True)          #removes all the junk after last name
df['Last Name'] = df['Last Name'].str.title()                                   #Capitalize first letter, lowercase rest
df['L Name'] = df['Last Name'].str.match("^[_A-z|(-|'|á|ã|é|ñ|ó|è|ç|ć|Ô|í)?]*((\s)*[_A-z]|(-|é|ñ|ó|è|ç|Ô|í)?)*$")

check_exists('Company Name', 'Comp Good', df)

#if city_check == "on":
if not df['City'].isna().all():     #checks to see if columns is empty
    df['City'] = df['City'].str.replace(r"^ +| +$|,+$", r"", regex=True)    #removes leading trailing spaces and comma at the end
    df['City'] = df['City'].str.title()                                   #Capitalize first letter, lowercase rest

# df['State'] = df['State'].replace(r"^([a-z])", '/\U/$1', regex=True)
#df.loc[df['Campaign ID'].apply(len) == 18, 'CID_status'] = 'TRUE'
#if df.loc[df['State'].str.len() > 2]:
#    df['State'] = df['State'].str.title()

if any(df['State'].apply(lambda x: len(str(x)) > 2)):    #if state is more than 2 char i.e. not abbreviated, capitalize first letters to normalize against map
    df['State'] = df['State'].str.title()

state_check('State', df)    #converts long state name to abbreviation
check_required_col('State','State Good',df,df_states,'Abbreviation')


#check zip code length for 5 characters or 5+4 or Canadian zip
search = SearchEngine()
df['Zip'] = df['Zip'].astype(str)
#df['Zip'] = df['Zip'].replace(r".0", r"", regex=True)
df['Zip'] = df['Zip'].apply(lambda x : str(x).zfill(5))
df['Zip Status'] = df['Zip'].str.match("^[']?[0-9]{5}(?:-[0-9]{4})?$|([ABCEGHJ-NPRSTVXY]\d[ABCEGHJ-NPRSTV-Z][ -]?\d[ABCEGHJ-NPRSTV-Z]\d)")


# for zip/state checking
search = SearchEngine()
df['City Lookup'] = df['Zip'].apply(zip_city)
df['State Lookup'] = df['Zip'].apply(zip_state)


check_required_col('Country','Country Good',df,df_acceptable,'Country')


if not df['Phone'].isna().all():        #check to see if column is empty, dont run the check
#df['Ph status'] = pd.np.where(df.Phone.isnull(), "optional")
    if df['Phone'].dtypes != 'object':
        df['Phone'] = df['Phone'].astype(str)
    df['Phone'] = df['Phone'].replace('nan', np.nan)
    df['Phone'] = df['Phone'].str.replace('\D', '', regex=True)
    df['Phone'].replace(to_replace="^[1]", value=r"", regex=True, inplace=True)  #remove the leading 1
    if phone_remove_zeros == 'on':
        df['Phone'] = df['Phone'].str.replace('0$', '', regex=True)  #if all entries have trailing zero
    df['Ph status'] = df['Phone'].str.match("[0-9]{10}$")
    df.loc[df['Phone'].isnull(), 'Ph status'] = 'optional'



#test fixes
df.loc[df['Phone'].isnull(), 'OPT IN PHONE'] = 'N'

#clean up the opt in responses - required - chnage this to check only
opt_in(df,'OPT IN EMAIL')
opt_in(df,'OPT IN MAIL')
opt_in(df,'OPT IN PHONE')
opt_in(df,'OPT IN THIRD PARTY')

check_required_col('Product Interest','Prod Int',df,df_acceptable, 'Product Interest')

check_optional_col('Additional Product Interest','Addl Prod Int',df,df_acceptable)


check_optional_col('Estimated Number of Units','Est Num Units',df,df_acceptable)

check_optional_col('Timeline for Purchasing','Timeline',df,df_acceptable)

check_required_col('Industry','Industry_good',df,df_acceptable, 'Industry')

check_optional_col('Functional Area/Department','Funct Area',df,df_acceptable)

check_optional_col('Job Function','Job Funct',df,df_acceptable)

# check employee range and fix if date is showing and then compare to acceptable values
check_required_col('Employee Range','emp range good',df,df_acceptable, 'Employee Range')

check_optional_col('Revenue Range','Revenue Rg',df,df_acceptable)

check_optional_col('Budget Established','Bud Est',df,df_acceptable)

df.loc[df['Request Follow-Up/Demo'] == 'YES', 'Request Follow-Up/Demo'] = 'Yes'
check_optional_col('Request Follow-Up/Demo','Req FU',df,df_acceptable)


#this looks for FALSE in the column and if it exists color the header red

for column in df.columns:
    if not df[column].all():
        name = column + "-error"
#        print(column + " has errors")
#        print(name)
        df = df.rename(columns= {column:name})


#add in Mobile Phone column if it doesn't exist - add in here so it doesn't get flagged with error from above check
if not 'Mobile Phone' in df.columns:
    idx=df.columns.get_loc("OPT IN EMAIL")
    #idx = idx-1
    df.insert(idx, "Mobile Phone", "")


#sum issues and warnings
issues = (df == False).sum().sum()
issues = issues + (df == 'required').sum().sum()
#warnings = (df == 'set OIP N').sum().sum()

print("There are " + str(issues) + " issues.")
#print("There are " + str(warnings) + " warnings.")

#save df file to temp excel doc for openpyxl
df.to_excel(working_folder + "\working_copy.xlsx", index=False)


#make a copy of df and find any columns with 'Field' in the name and create a df and drop if empty
#how to use this??
df_field = df.copy()
df_field = df_field.filter(like='Field', axis=1)
df_field.dropna(how='all', axis=1, inplace=True)
#df_field1 =pd.concat([df, df_field], axis='columns')
#print(df_field)


#openpyxl conditional formatting section
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

wb = load_workbook(working_folder + "\working_copy.xlsx")
ws = wb.active

red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(fill=red_fill)
yellow_fill = PatternFill(bgColor="FFFF00")
dxy = DifferentialStyle(fill=yellow_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("FALSE",N2)))']
rule1 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule1.formula = ['NOT(ISERROR(SEARCH("required",N2)))']
rule2 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxy)
rule2.formula = ['NOT(ISERROR(SEARCH("set OIP N",N2)))']
rule3 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule3.formula = ['NOT(ISERROR(SEARCH("error",A1)))']
ws.conditional_formatting.add('N2:BT4000', rule)
ws.conditional_formatting.add('N2:BT4000', rule1)
ws.conditional_formatting.add('N2:BT4000', rule2)
ws.conditional_formatting.add('A1:DA1', rule3)


wb.save(working_folder + "\working_copy.xlsx")
