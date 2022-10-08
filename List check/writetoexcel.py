import openpyxl
import pandas

#now lets copy the columns from raw to the template but I only know how to do it in pandas
df_template = openpyxl.load_workbook(r'c:\test\new.xlsx')
df_filtered = openpyxl.load_workbook(r'c:\test\testrawdata_filtered.xlsx')

wb3 = openpyxl.load_workbook(r'c:\test\col_map.xlsx')
ws3 = wb3['Sheet1']

x = (ws3.cell(3,2).value)
print(x)


from openpyxl.utils import rows_from_range

def copy_range(range_str, src, dst):

    for row in rows_from_range(range_str):
        for cell in row:
            dst[cell].value = src[cell].value

    return


sheet = df_template['Template (File Format)']
sheet1 = df_filtered['Sheet1']
for row in sheet1['A2':'Z100']:
    for cell in row:
        sheet[cell.coordinate].value = cell.value


#print (df_template)
df_template.save(r'C:\test\testrawdata_final.xlsx')



