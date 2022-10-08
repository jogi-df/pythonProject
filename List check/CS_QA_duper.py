from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd


df = pd.read_excel(r'C:\test\CS requestor.xlsx')
df['sheet'] = df['Prod Code'] + '-' + df['CID']

names = df['sheet'].to_list()
names1 = df['Program Name'].to_list()
cid = df['CID'].to_list()
pid = df['ParentID'].to_list()

wb = load_workbook(r"C:\test\Adobe Sign - Content Syndication - QA Grid Shell.xlsx")
ws = wb['Child Program (dupe as needed)']

#i=1 needed?

#create new worksheets based on <prod code>-<CID> and pre populate fields
for i in range(len(names)):
    wb.copy_worksheet(ws)
    ws = wb.worksheets[-1]
    ws.title = names[i]
    #sheet = wb.active
    ws.cell(row=15, column=2).value = names1[i]
    ws.cell(row=2, column=2).value = 'NA_FY[#]_[Q#]_[FUNNEL]_[CS/Social/DIS/SEM/Web]_[WP/eBook/Blog/Cstudy/Infogr]_[ContentName]_[VendorName]_[CCT/Sign]_[SFDC Campaign ID]'
    ws.cell(row=6, column=2).value = 'Anne Wong'
    ws.cell(row=13, column=1).value = 'Folder to Clone to: 22Q4 Substance Content Syndication'

#add CID and token ID into Tokens sheet
for j in range(len(cid)):
    k = j + j + 6
    m = j + j + 7

    ws = wb['Tokens']
    ws.cell(row=k, column=1).value = '{{my.ParentID}}'
    ws.cell(row=k, column=2).value = pid[j]
    ws.cell(row=m, column=1).value = '{{my.SFDC Campaign ID}}'
    ws.cell(row=m, column=2).value = cid[j]

ws = wb['Tokens']
for row in ws[2:ws.max_row]:  # skip the header
    cell = row[0]
    cell1 = row[1]
    cell.alignment = Alignment(horizontal='left')
    cell1.alignment = Alignment(horizontal='left')

startRow = 1
endRow = 2 * len(cid) + 6
for i in range(startRow,endRow):
    ws.row_dimensions[i+1].height=28.5

wb.save(r"C:\test\Adobe Sign - Content Syndication - QA Grid Shell-copy.xlsx")