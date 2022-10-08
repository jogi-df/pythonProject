import openpyxl
import pandas

wb1 = openpyxl.load_workbook(r'c:\test\testrawdata_filtered.xlsx')
ws1 = wb1['Sheet1']

wb2 = openpyxl.load_workbook(r'c:\test\new.xlsx')
ws2 = wb2['Template (File Format)']

wb3 = openpyxl.load_workbook(r'c:\test\col_map.xlsx')
ws3 = wb3['Sheet1']


## Create a dictionary of column names for filtered data
ColNames1 = {}
Current1  = 0
for COL in ws1.iter_cols(1, ws1.max_column):
    ColNames1[COL[0].value] = Current1
    Current1 += 1

## Create a dictionary of column names for template (file format)
ColNames2 = {}
Current2  = 0
for COL in ws2.iter_cols(1, ws2.max_column):
    ColNames2[COL[0].value] = Current2
    Current2 += 1

print(ColNames1)

#this should print 0
#print(ColNames2['Campaign ID'])

x = (ws3.cell(2,1).value)
print(x)


#print(ColNames2['Campaign ID'])
## create a dataframe from maps for column matching lookups
#print(dfm.iloc[0]) #this should print






