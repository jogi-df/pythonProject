#create new workbook and rename the default sheet to Template (File Format)
from openpyxl import load_workbook, Workbook
from copy import copy

original_wb = load_workbook(r"C:\test\Lead Import Template Worldwide.xlsx")
original_sheet = original_wb['Template (File Format)']


new_wb = Workbook()
new_ws=new_wb['Sheet']
new_ws.title ='Template (File Format)'

#copy over column width
for z, cd in original_sheet.column_dimensions.items():
    new_ws.column_dimensions[z].width = cd.width

#copy over all other formatting
for row in original_sheet.rows:
    for cell in row:
        new_ws[cell.coordinate] = cell.value
        new_cell = new_ws[cell.coordinate]

        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


new_wb.save(r'c:\test\new.xlsx')
