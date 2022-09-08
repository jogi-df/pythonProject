from flask import Flask, render_template, request, session, send_from_directory, send_file
import pandas as pd
import os
from werkzeug.utils import secure_filename
from functions import check_optional_col, check_required_col, opt_in, check_exists, check_phone_col
from validate_email import validate_email

#*** Flask configuration

# Define folder to save uploaded files to process further
path = '/home/jogidf/mysite'
UPLOAD_FOLDER = os.path.join(path, 'static', 'uploads')

# Define allowed files (for this example I want only csv file)
ALLOWED_EXTENSIONS = {'csv', 'xls', 'xlsx'}

app = Flask(__name__, template_folder='templates', static_folder='static')
# Configure upload file path flask
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Define secret key to enable session
app.secret_key = '34534lkj523l4kj;l342gf-'


@app.route('/')
def index():
    return render_template('menu.html')


@app.route('/input')
def index1():
    return render_template('index_upload_and_show_data.html')

@app.route('/input',  methods=("POST", "GET"))
def uploadFile():
    if request.method == 'POST':
        # upload file flask - first file should be the template, second is the file to be analyzed
        uploaded_df = request.files['uploaded-file']
        uploaded_df2 = request.files['uploaded-file2']

        # Extracting uploaded data file name
        data_filename = secure_filename(uploaded_df.filename)
        data_filename2 = secure_filename(uploaded_df2.filename)

        # flask upload file to database (defined uploaded folder in static path)
        uploaded_df.save(os.path.join(app.config['UPLOAD_FOLDER'], data_filename))
        uploaded_df2.save(os.path.join(app.config['UPLOAD_FOLDER'], data_filename2))

        # Storing uploaded file path in flask session
        session['uploaded_data_file_path'] = os.path.join(app.config['UPLOAD_FOLDER'], data_filename)
        session['uploaded_data_file_path2'] = os.path.join(app.config['UPLOAD_FOLDER'], data_filename2)

        return render_template('index_upload_and_show_data_page2.html')

@app.route('/show_data')
def showData():
    # Retrieving uploaded file path from session
    data_file_path = session.get('uploaded_data_file_path', None)
    data_file_path2 = session.get('uploaded_data_file_path2', None)

    # read csv file in python flask (reading uploaded csv file from uploaded server location)
    uploaded_df = pd.read_excel(data_file_path2, engine='openpyxl')


    #set new dataframes to pass to the check section
    df = uploaded_df
    df_acceptable = pd.read_excel(data_file_path, sheet_name="Acceptable List of Values", engine='openpyxl')
    df_states = pd.read_excel(data_file_path, sheet_name="Territory State List", engine='openpyxl')

    #-----------start check section---------------#

    #remove any trailing whitespaces in the column names (shows up in permissions create date)
    df.columns = df.columns.str.rstrip()

    #sometimes raw data comes in on the template where Campaign ID is the CID column.  Lets normalize the name.
    if 'Campaign ID' in df.columns:
        df.rename(columns = {"Campaign ID":"CID"}, inplace = True)


    #check CID length for 18 characters
    df.loc[df['CID'].apply(len) == 18, 'CID_status'] = 'TRUE'
    df.loc[df['CID'].apply(len) != 18, 'CID_status'] = 'FALSE'

    #permissions date cleanup
    df['Permissions Create Date'] = df['Permissions Create Date'].astype(str)
    df['Perm Date'] = df['Permissions Create Date'].str.match("[0-9]{8}")

    check_optional_col('Attended','Attend',df,df_acceptable)

    #check if email address is in a standard format
    df['Email_Good'] = df['Email'].apply(validate_email)

    check_optional_col('Salutation','Sal Good',df,df_acceptable)

    check_exists('First Name', 'F Name',df)

    check_exists('Last Name', 'L Name', df)

    check_exists('Company Name', 'Comp Good', df)

    check_required_col('State','State Good',df,df_states,'Abbreviation')

    #check zip code length for 5 characters or 5+4
    df['Zip'] = df['Zip'].astype(str)
    df['Zip Status'] = df['Zip'].str.match("^[0-9]{5}(?:-[0-9]{4})?$")

    check_required_col('Country','Country Good',df,df_acceptable,'Country')

    check_phone_col('Phone', 'Ph Status', df)

    #clean up the opt in responses - required
    opt_in(df,'OPT IN EMAIL')
    opt_in(df,'OPT IN MAIL')
    opt_in(df,'OPT IN PHONE')
    opt_in(df,'OPT IN THIRD PARTY')

    check_required_col('Product Interest','Prod Int',df,df_acceptable, 'Product Interest')

    check_optional_col('Additional Product Interest','Addl Prod Int',df,df_acceptable)

    check_optional_col('Estimated Number of Units','Est Num Units',df,df_acceptable)

    check_optional_col('Timeline for Purchasing','Timeline',df,df_acceptable)

    check_required_col('Industry','Industry_good',df,df_acceptable, 'Industry')

    check_optional_col('Functional Area/Department','Funct Area',df,df_acceptable)

    check_optional_col('Job Function','Job Funct',df,df_acceptable)

    check_required_col('Employee Range','emp range good',df,df_acceptable, 'Employee Range')

    check_optional_col('Revenue Range','Revenue Rg',df,df_acceptable)

    check_optional_col('Budget Established','Bud Est',df,df_acceptable)

    check_optional_col('Request Follow-Up/Demo','Req FU',df,df_acceptable)

    #---------------------end check section------------------#

    #save df file to temp excel doc for openpyxl
    df.to_excel(UPLOAD_FOLDER +"/" + "working_copy.xlsx", index=False)

    #make a copy of df and find any columns with 'Field' in the name and create a df and drop if empty
    #how to use this??
    df_field = df.copy()
    df_field = df_field.filter(like='Field', axis=1)
    df_field.dropna(how='all', axis=1, inplace=True)
    #df_field1 =pd.concat([df, df_field], axis='columns')
    #print(df_field)


    #openpyxl conditional formatting section
    from openpyxl import load_workbook
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting import Rule

    wb = load_workbook(UPLOAD_FOLDER +"/" + "working_copy.xlsx")
    ws = wb.active

    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(fill=red_fill)
    yellow_fill = PatternFill(bgColor="FFFF00")
    dxy = DifferentialStyle(fill=yellow_fill)
    rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("FALSE",AP2)))']
    rule1 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
    rule1.formula = ['NOT(ISERROR(SEARCH("required",AP2)))']
    rule2 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxy)
    rule2.formula = ['NOT(ISERROR(SEARCH("set OIP N",AP2)))']
    ws.conditional_formatting.add('AP2:BT4000', rule)
    ws.conditional_formatting.add('AP2:BT4000', rule1)
    ws.conditional_formatting.add('AP2:BT4000', rule2)
    wb.save(UPLOAD_FOLDER +"/" + "working_copy.xlsx")

    #download the working copy file
    #app.config['UPLOAD_FOLDER'] = "/home/jogidf/mysite/static/uploads/"

    filename ="/home/jogidf/mysite/static/uploads/working_copy.xlsx"

    issues = (df == False).sum().sum()
    issues = issues + (df == 'required').sum().sum()
    warnings = (df == 'set OIP N').sum().sum()

    return render_template('result.html', issues=issues, warnings=warnings)

    #return send_file(filename, as_attachment=True)

    #return render_template('fix_file.html')

    # pandas dataframe to html table flask
    #uploaded_df_html = uploaded_df.to_html()
    #return render_template('show_csv_data.html', data_var = uploaded_df_html)

@app.route('/fix')
def fix():
    return render_template('fix_file.html')
    #return render_template('index_upload_and_show_data.html')

@app.route('/fix',  methods=("POST", "GET"))
def uploadFix():
    if request.method == 'POST':
        # upload file flask - first file should be the template, second is the file to be analyzed
        uploaded_fix = request.files['fixfile1']

        # Extracting uploaded data file name
        data_filename3 = secure_filename(uploaded_fix.filename)

        # flask upload file to database (defined uploaded folder in static path)
        uploaded_fix.save(os.path.join(app.config['UPLOAD_FOLDER'], data_filename3))

        # Storing uploaded file path in flask session
        session['uploaded_data_file_path3'] = os.path.join(app.config['UPLOAD_FOLDER'], data_filename3)

        # store checkbox values
        session['checkbox1'] = request.form.get('pmdate')
        session['checkbox2'] = request.form.get('zcode')
        session['checkbox3'] = request.form.get('optin')
        session['checkbox4'] = request.form.get('optinu')
        session['checkbox5'] = request.form.get('emprange')
        session['checkbox6'] = request.form.get('final')

        return render_template('fix_file2.html')

@app.route('/results')
def results():
    data_file_path3 = session.get('uploaded_data_file_path3', None)
    cb1 = session.get('checkbox1', None)
    cb2 = session.get('checkbox2', None)
    cb3 = session.get('checkbox3', None)
    cb4 = session.get('checkbox4', None)
    cb5 = session.get('checkbox5', None)
    cb6 = session.get('checkbox6', None)

    # read csv file in python flask (reading uploaded csv file from uploaded server location)
    uploaded_df3 = pd.read_excel(data_file_path3, engine='openpyxl')


    #set new dataframes to pass to the check section
    dfr = uploaded_df3

    if cb1 == 'on':
        dfr["Permissions Create Date"] = pd.to_datetime(dfr["Permissions Create Date"]).dt.strftime("%m%d%Y")

    if cb2 == 'on':
        dfr['Zip'] = dfr['Zip'].apply(lambda x : str(x).zfill(5))

    if cb3 == 'on':
        opt_in(dfr,'OPT IN EMAIL')
        opt_in(dfr,'OPT IN MAIL')
        opt_in(dfr,'OPT IN PHONE')
        opt_in(dfr,'OPT IN THIRD PARTY')

    if cb4 == 'on':
        dfr.loc[dfr['Phone'].isnull(), 'OPT IN PHONE'] = 'N'

    if cb5 == 'on':
        dfr.loc[dfr['Employee Range'] == 'Oct-99', 'Employee Range'] = '10-99'
        dfr.loc[dfr['Employee Range'] == '9-Jan', 'Employee Range'] = '1-9'

    dfr.to_excel(UPLOAD_FOLDER +"/" + "working_copy.xlsx", index=False)
    return render_template('temp.html', cb1=cb1)



if __name__=='__main__':
    app.run(debug = True)


