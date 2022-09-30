import pandas as pd
import easygui
import openpyxl as op
import numpy as np
import os
from functions import check_optional_col, check_required_col, opt_in, opt_in_check, check_phone_col, whitespace_remover, check_exists
from validate_email import validate_email
from io import StringIO
import phonenumbers

#reference files
#df_zip = pd.read_excel(r'C:\test\ZIP_Locale_Detail.xls')



#sets the working folder
working_folder = easygui.diropenbox(msg="Select the working folder", title="Working folder")
fold = StringIO(working_folder)


#easygui requests a file select for the main data frame.  Easygui returns the path of the file
raw_file_path = easygui.fileopenbox(msg="Select the raw data file.  Don't use the original!", default=working_folder + "\*")
file = StringIO(raw_file_path)
df = pd.read_excel(raw_file_path)
#df = pd.read_excel(r"C:\test\testrawdata.xlsx")   #uncomment this for testing and comment out everything else

# these are from the Adobe Lead Import Template Worldwide tabs and are used to compare against
#from io import StringIO
#raw_file_path1 = easygui.fileopenbox(msg="Select the template data file.  Don't use the original!")
#file = StringIO(raw_file_path1)
#df_acceptable = pd.read_excel(raw_file_path1, sheet_name="Acceptable List of Values")
#df_states = pd.read_excel(raw_file_path1, sheet_name="Territory State List")
df_acceptable = pd.read_excel(r'C:\test\Lead Import Template Worldwide.xlsx', sheet_name="Acceptable List of Values")          #uncomment this for testing and comment out everything else
df_states = pd.read_excel(r'C:\test\Lead Import Template Worldwide.xlsx', sheet_name="Territory State List")                   #uncomment this for testing and comment out everything else

#remove any trailing whitespaces in the column names (shows up in permissions create date)
df.columns = df.columns.str.rstrip()

#whitespace_remover(df)

#sometimes raw data comes in on the template where Campaign ID is the CID column.  Lets normalize the name.
#if 'Campaign ID' in df.columns:
#    df.rename(columns = {"Campaign ID":"CID"}, inplace = True)

#if 'Zip Code' in df.columns:
#    df.rename(columns = {"Zip Code":"Zip"}, inplace = True)

df['CID'] = df.filter(like="Campaign")

df['State'] = df.filter(like="State")

df['Zip'] = df.filter(like="Zip")

df['Email'] = df.filter(like="Email")

df['Country'] = df.filter(like="Country")

df['Company Name'] = df.filter(like="Company")


#check CID length for 18 characters
df.loc[df['CID'].apply(len) == 18, 'CID_status'] = 'TRUE'
df.loc[df['CID'].apply(len) != 18, 'CID_status'] = 'FALSE'


#df['Perm Date'] = df['Permissions Create Date'].str.match("\'[0-9]{8}")

df['Permissions Create Date'] = df['Permissions Create Date'].astype(str)

#df['Permissions Create Date'] = df['Permissions Create Date'].str.zfill(8)

print(df['Permissions Create Date'])

#df.loc[df["Permissions Create Date"] != df["Permissions Create Date"].str.match("^[']?[0-1]?[1-9][1-3][1-9](202)[2-3]"), df["Permissions Create Date"]] = pd.to_datetime(df["Permissions Create Date"]).dt.strftime("%m%d%Y")

#df.loc[df['Zip'].str.match("^[0-9]{4}") == True, 'Zip'] = df.loc['Zip'].str.zfill(5)

df['perm'] = df["Permissions Create Date"].str.match("^[1-9][1-3][1-9](202)[2-3]")

#print(df['perm'])

df.loc[df["Permissions Create Date"].str.match("^[1-9][1-3][1-9](202)[2-3]") == 'True', 'Permissions Create Date'] = df['Permissions Create Date'].str.zfill(8)
#df.loc[df[) == 'True', 'perm'] = df['Permissions Create Date'].str.zfill(8)

print(df['Permissions Create Date'])

#df.loc[df["Permissions Create Date"].str.match("^[1-9][1-3][1-9](202)[2-3]") == 'TRUE', 'perm'] = df['Permissions Create Date'].str.zfill(10)

print(df['perm'])

#df["Permissions Create Date"] = pd.to_datetime(df["Permissions Create Date"]).dt.strftime("%m%d%Y")




#if not "Permissions Create Date" in df.columns:
#    print(colchk + " column not in source")
#        return
#
#    df["Permissions Create Date"] = df["Permissions Create Date"].astype(str)
#    datasrc[colres] = datasrc[colchk].str.match("^[']?[0-1][0-9][1-3][1-9](202)[2-3]")






check_optional_col('Attended','Attend',df,df_acceptable)

#check if email address is in a standard format
df['Email_Good'] = df['Email'].apply(validate_email)

check_optional_col('Salutation','Sal Good',df,df_acceptable)

check_exists('First Name', 'F Name',df)

check_exists('Last Name', 'L Name', df)

check_exists('Company Name', 'Comp Good', df)

check_required_col('State','State Good',df,df_states,'Abbreviation')

#check zip code length for 5 characters or 5+4
df['Zip'] = df['Zip'].astype(str)
df['Zip Status'] = df['Zip'].str.match("^[0-9]{5}(?:-[0-9]{4})?$")
#df.loc[df['Zip'].str.match("^[0-9]{4}") == True, 'Zip'] = df.loc['Zip'].str.zfill(5)
#[df['Zip'].str.match("^[0-9]{4}") == True, df['Zip']] = df['Zip'].str.zfill(5)


#df['Perm Date'] = df['Permissions Create Date'].str.match("[0-9]{8}")

check_required_col('Country','Country Good',df,df_acceptable,'Country')

check_phone_col('Phone', 'Ph Status', df)

#test fixes
df.loc[df['Phone'].isnull(), 'OPT IN PHONE'] = 'N'

df['Zip'] = df['Zip'].apply(lambda x : str(x).zfill(5))


#df.loc[df['Employee Range'] == '10/1/1999', 'Employee Range'] = '\'10-99'
#df.loc[df['Employee Range'] == '1/9/2022', 'Employee Range'] = '\'1-9'


#clean up the opt in responses - required - chnage this to check only
opt_in_check(df,'OPT IN EMAIL')
opt_in_check(df,'OPT IN MAIL')
opt_in_check(df,'OPT IN PHONE')
opt_in_check(df,'OPT IN THIRD PARTY')

check_required_col('Product Interest','Prod Int',df,df_acceptable, 'Product Interest')

check_optional_col('Additional Product Interest','Addl Prod Int',df,df_acceptable)

check_optional_col('Estimated Number of Units','Est Num Units',df,df_acceptable)

check_optional_col('Timeline for Purchasing','Timeline',df,df_acceptable)

check_required_col('Industry','Industry_good',df,df_acceptable, 'Industry')

check_optional_col('Functional Area/Department','Funct Area',df,df_acceptable)

check_optional_col('Job Function','Job Funct',df,df_acceptable)

# check employee range and fix if date is showing and then compare to acceptable values
check_required_col('Employee Range','emp range good',df,df_acceptable, 'Employee Range')

check_optional_col('Revenue Range','Revenue Rg',df,df_acceptable)

check_optional_col('Budget Established','Bud Est',df,df_acceptable)

check_optional_col('Request Follow-Up/Demo','Req FU',df,df_acceptable)

issues = (df == False).sum().sum()

issues = issues + (df == 'required').sum().sum()
warnings = (df == 'set OIP N').sum().sum()

print(issues)

#save df file to temp excel doc for openpyxl
df.to_excel(working_folder + "\working_copy.xlsx", index=False)


#make a copy of df and find any columns with 'Field' in the name and create a df and drop if empty
#how to use this??
df_field = df.copy()
df_field = df_field.filter(like='Field', axis=1)
df_field.dropna(how='all', axis=1, inplace=True)
#df_field1 =pd.concat([df, df_field], axis='columns')
#print(df_field)


#openpyxl conditional formatting section
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

wb = load_workbook(working_folder + "\working_copy.xlsx")
ws = wb.active

red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(fill=red_fill)
yellow_fill = PatternFill(bgColor="FFFF00")
dxy = DifferentialStyle(fill=yellow_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("FALSE",N2)))']
rule1 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule1.formula = ['NOT(ISERROR(SEARCH("required",N2)))']
rule2 = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxy)
rule2.formula = ['NOT(ISERROR(SEARCH("set OIP N",N2)))']
ws.conditional_formatting.add('N2:BT4000', rule)
ws.conditional_formatting.add('N2:BT4000', rule1)
ws.conditional_formatting.add('N2:BT4000', rule2)
wb.save(working_folder + "\working_copy.xlsx")
